"""Excel report generation."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from cube_budget.core.models import OptimizationResult
from cube_budget.reports.summary import SummaryStats


class ExcelReporter:
    """Generate Excel reports with multiple sheets."""

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def generate(
        self,
        result: OptimizationResult,
        output_dir: str | Path,
        recent_runs: list[dict] | None = None,
    ) -> Path:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filepath = output_dir / f"cube_budget_report_{timestamp}.xlsx"

        stats = SummaryStats.compute(result)
        wb = Workbook()

        self._write_summary(wb.active, result, stats)
        self._write_by_store(wb.create_sheet("Por Loja"), result, stats)
        self._write_assignments(wb.create_sheet("Atribuições"), result)
        self._write_missing(wb.create_sheet("Não Encontradas"), result)
        self._write_analysis(wb.create_sheet("Análise"), result, stats)
        if recent_runs:
            self._write_history(wb.create_sheet("Histórico"), recent_runs)

        wb.save(filepath)
        return filepath

    def _style_header(self, ws, row: int, cols: int) -> None:
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

    def _write_summary(self, ws, result: OptimizationResult, stats: dict) -> None:
        ws.title = "Resumo"
        rows = [
            ("Métrica", "Valor"),
            ("Total de cartas", result.total_cards),
            ("Cartas encontradas", f"{result.found_cards} ({stats['found_pct']:.1f}%)"),
            ("Cartas não encontradas", len(result.missing)),
            ("Lojas necessárias", result.stores_used),
            ("Preço total", f"R$ {result.total_price:,.2f}"),
            ("Preço médio", f"R$ {stats['avg_price']:,.2f}"),
            ("Solver utilizado", result.solver),
            ("Tempo de otimização", f"{result.duration_ms / 1000:.1f}s"),
            ("Greedy (benchmark)", f"{result.greedy_stores} lojas, R$ {result.greedy_price or 0:,.2f}"),
            ("Economia vs Greedy", f"R$ {stats['savings']:,.2f}"),
            ("Run UUID", result.run_uuid),
        ]
        for i, (label, value) in enumerate(rows, 1):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
        self._style_header(ws, 1, 2)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _write_by_store(self, ws, result: OptimizationResult, stats: dict) -> None:
        headers = ["Loja", "Cartas", "Subtotal"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        row = 2
        for store_name in sorted(stats["store_totals"].keys()):
            ws.cell(row=row, column=1, value=store_name)
            ws.cell(row=row, column=2, value=stats["store_counts"][store_name])
            ws.cell(row=row, column=3, value=f"R$ {stats['store_totals'][store_name]:,.2f}")
            row += 1

    def _write_assignments(self, ws, result: OptimizationResult) -> None:
        headers = ["Carta", "Loja", "Condição", "Idioma", "Edição", "Preço"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        for i, a in enumerate(result.assignments, 2):
            ws.cell(row=i, column=1, value=a.card.raw_name)
            ws.cell(row=i, column=2, value=a.store.name)
            ws.cell(row=i, column=3, value=a.offer.condition)
            ws.cell(row=i, column=4, value=a.offer.language)
            ws.cell(row=i, column=5, value=a.offer.edition or "")
            ws.cell(row=i, column=6, value=a.price)

    def _write_missing(self, ws, result: OptimizationResult) -> None:
        headers = ["Carta", "Motivo"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        for i, m in enumerate(result.missing, 2):
            ws.cell(row=i, column=1, value=m.raw_name)
            ws.cell(row=i, column=2, value=m.reason)

    def _write_analysis(self, ws, result: OptimizationResult, stats: dict) -> None:
        ws.cell(row=1, column=1, value="Top 10 Cartas Mais Caras")
        self._style_header(ws, 1, 3)
        ws.cell(row=1, column=2, value="Loja")
        ws.cell(row=1, column=3, value="Preço")

        for i, a in enumerate(stats["top_expensive"], 2):
            ws.cell(row=i, column=1, value=a.card.raw_name)
            ws.cell(row=i, column=2, value=a.store.name)
            ws.cell(row=i, column=3, value=f"R$ {a.price:,.2f}")

    def _write_history(self, ws, recent_runs: list[dict]) -> None:
        headers = ["Data", "Cartas", "Lojas", "Preço", "Solver"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header(ws, 1, len(headers))

        for i, run in enumerate(recent_runs, 2):
            ws.cell(row=i, column=1, value=run.get("created_at", ""))
            ws.cell(row=i, column=2, value=run.get("total_cards", ""))
            ws.cell(row=i, column=3, value=run.get("stores_count", ""))
            ws.cell(row=i, column=4, value=run.get("total_price", ""))
            ws.cell(row=i, column=5, value=run.get("solver_used", ""))
